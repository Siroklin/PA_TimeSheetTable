import calendar
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

WORK_COLOR = 'D4EDDA'
SHIFT_COLORS = {
    'В':  'FFF3CD',
    'О':  'CCE5FF',
    'Б':  'F8D7DA',
    'ДО': 'FDE8C8',
    'П':  'FF8787',
    'К':  '99E9F2',
    'Ф':  'E8D5F5',
    'У':  'FF4D4F',
    'Д':  'FFD43B',
    'С':  'FFA94D',
}
DAY_EMPTY = 'EBEBEB'
NIGHT_EMPTY = 'C8C8C8'

LEGEND = [
    ('8',   WORK_COLOR, 'Работает (часы по графику)'),
    ('В',   'FFF3CD', 'Выходной'),
    ('О',   'CCE5FF', 'Отпуск'),
    ('Б',   'F8D7DA', 'Больничный'),
    ('ДО',  'FDE8C8', 'Отпуск за свой счёт'),
    ('П',   'FF8787', 'Прогул'),
    ('К',   '99E9F2', 'Командировка'),
    ('Ф',   'E8D5F5', 'ФМС'),
    ('У',   'FF4D4F', 'Увольнение'),
    ('Д',   'FFD43B', 'Доп. смена'),
    ('С',   'FFA94D', 'Сверхурочные'),
]

_CODE_RE = re.compile(r'^(\D*)(\d+)?$')


def _code_of(value: str) -> str:
    if not value:
        return ''
    m = _CODE_RE.match(value)
    return (m.group(1) if m else value) or ''


def _cell_color(value: str, empty: str) -> str:
    """Splits a value like 'Д4' or '8' into a letter code + hour count and
    returns the matching fill color (work hours have no letter prefix)."""
    if not value:
        return empty
    code = _code_of(value)
    if code == '':
        return WORK_COLOR
    return SHIFT_COLORS.get(code, empty)

MONTH_NAMES = [
    'Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь',
    'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь',
]

_thin = Side(style='thin', color='AAAAAA')
_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill(fill_type='solid', fgColor=hex_color)


def generate_excel(employees, schedule_map: dict, year: int, month: int, department: str) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "График"

    days = calendar.monthrange(year, month)[1]
    total_cols = 3 + days * 2

    # Row 1: Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title = ws.cell(row=1, column=1, value=f"График работы — {department}, {MONTH_NAMES[month - 1]} {year}")
    title.font = Font(bold=True, size=13)
    title.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 26

    # Row 2: fixed headers + merged day numbers
    ws.cell(row=2, column=1, value='№')
    ws.cell(row=2, column=2, value='ФИО')
    ws.cell(row=2, column=3, value='Должность')
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 26
    ws.column_dimensions['C'].width = 16

    for d in range(1, days + 1):
        col_d = 4 + (d - 1) * 2
        col_n = col_d + 1
        ws.merge_cells(start_row=2, start_column=col_d, end_row=2, end_column=col_n)
        c = ws.cell(row=2, column=col_d, value=str(d))
        c.font = Font(bold=True, size=9)
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[get_column_letter(col_d)].width = 3.5
        ws.column_dimensions[get_column_letter(col_n)].width = 3.5

    # Row 3: Д/Н sub-headers
    for d in range(1, days + 1):
        col_d = 4 + (d - 1) * 2
        ws.cell(row=3, column=col_d, value='Д').alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=3, column=col_d + 1, value='Н').alignment = Alignment(horizontal='center', vertical='center')

    _hfont = Font(bold=True, size=9)
    for row in (2, 3):
        for col in range(1, total_cols + 1):
            c = ws.cell(row=row, column=col)
            c.font = _hfont
            c.border = _BORDER
            c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 16

    # Data rows
    for i, emp in enumerate(employees, start=1):
        row = 3 + i
        ws.cell(row=row, column=1, value=i).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=row, column=2, value=emp.name)
        ws.cell(row=row, column=3, value=emp.position).alignment = Alignment(horizontal='center', vertical='center')

        emp_sched = schedule_map.get(emp.id, {})

        for d in range(1, days + 1):
            col_d = 4 + (d - 1) * 2
            # keys may be int or str depending on source
            cell_data = emp_sched.get(d) or emp_sched.get(str(d)) or {}
            day_val = cell_data.get('day') or ''
            night_val = cell_data.get('nightShift') or ''
            is_terminated = _code_of(day_val) == 'У' or _code_of(night_val) == 'У'
            terminated_color = SHIFT_COLORS['У']

            c_d = ws.cell(row=row, column=col_d, value=day_val)
            c_d.fill = _fill(terminated_color if is_terminated else _cell_color(day_val, DAY_EMPTY))
            c_d.alignment = Alignment(horizontal='center', vertical='center')
            c_d.font = Font(size=9)

            c_n = ws.cell(row=row, column=col_d + 1, value=night_val)
            c_n.fill = _fill(terminated_color if is_terminated else _cell_color(night_val, NIGHT_EMPTY))
            c_n.alignment = Alignment(horizontal='center', vertical='center')
            c_n.font = Font(size=9)

        for col in range(1, total_cols + 1):
            ws.cell(row=row, column=col).border = _BORDER
        ws.row_dimensions[row].height = 18

    # Legend
    legend_start = 3 + len(employees) + 2
    ws.cell(row=legend_start, column=1, value='Легенда:').font = Font(bold=True, size=9)
    ws.row_dimensions[legend_start].height = 14

    for offset, (letter, color, description) in enumerate(LEGEND, start=1):
        r = legend_start + offset
        cell = ws.cell(row=r, column=1, value=letter)
        cell.fill = _fill(color)
        cell.font = Font(bold=True, size=9)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = _BORDER
        ws.cell(row=r, column=2, value=description).font = Font(size=9)
        ws.row_dimensions[r].height = 16

    return wb
